from datetime import datetime, timedelta
import json
import os
import random
import string
from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from .forms import RegisterForm, KaizenSheetForm, HandwrittenKaizenForm
from .models import HorizontalDeployment, KaizenSheet, PasswordResetRequest, Profile, Image
from django.http import (
    HttpResponseBadRequest,
    JsonResponse,
    HttpResponse,
    HttpResponseForbidden,
)
from django.contrib import messages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage  # Rename to avoid confusion
from openpyxl.utils import range_boundaries, get_column_letter
from django.utils import timezone
from django.db import IntegrityError, models, transaction
from django.views.decorators.csrf import ensure_csrf_cookie
from .models import KaizenCoordinator
from django.views.decorators.http import require_POST
from .decorators import profile_required


# Home view
def index(request):
    if User.objects.exists():
        return redirect("login")
    else:
        return redirect("register")


def register(request):
    if request.method == "POST":
        # Get form data
        user_id = request.POST.get("employee_id")  
        username = request.POST.get("username")
        password = request.POST.get("password")
        email = request.POST.get("email")
        role = request.POST.get("user_type")  
        department = request.POST.get("department")

        # Validate required fields
        if not all([user_id, username, password, role]):
            missing_fields = []
            if not user_id:
                missing_fields.append("Employee ID")
            if not username:
                missing_fields.append("Username") 
            if not password:
                missing_fields.append("Password")
            if not role:
                missing_fields.append("Role")
            return JsonResponse(
                {"success": False, "error": f"Required fields missing: {', '.join(missing_fields)}"}
            )

        try:
            with transaction.atomic():
                # Check if employee ID is already used for this role
                if Profile.objects.filter(employee_id=user_id, user_type=role).exists():
                    return JsonResponse(
                        {"success": False, "error": f"Employee ID already exists for {role} role"}
                    )

                # Check HOD department uniqueness
                if role == "hod" and Profile.objects.filter(
                    user_type="hod", department=department
                ).exists():
                    return JsonResponse(
                        {"success": False, "error": f"HOD already exists for {department} department"}
                    )

                # Check coordinator limit
                if role == "coordinator":
                    if Profile.objects.filter(user_type="coordinator").count() >= 2:
                        return JsonResponse(
                            {"success": False, "error": "Maximum of 2 coordinators allowed"}
                        )

                if role == "hod_coordinator":
                    if not department:
                        return JsonResponse({
                            "success": False,
                            "error": "Department is required for HOD Coordinator"
                        })
                    
                    # Check if HOD Coordinator already exists for this department
                    if Profile.objects.filter(user_type="hod_coordinator", department=department).exists():
                        return JsonResponse({
                            "success": False,
                            "error": f"HOD Coordinator already exists for {department} department"
                        })


                # Check finance limit
                if role == "finance":
                    if Profile.objects.filter(user_type="finance").exists():
                        return JsonResponse(
                            {"success": False, "error": "Finance account already exists"}
                        )

                # Get or create user
                try:
                    user = User.objects.get(username=username)
                    # Check if user already has this role
                    if Profile.objects.filter(user=user, user_type=role).exists():
                        return JsonResponse(
                            {"success": False, "error": f"User already has {role} role"}
                        )
                except User.DoesNotExist:
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        email=email
                    )

                # Create new profile with proper department setting
                Profile.objects.create(
                    user=user,
                    user_type=role,
                    department=department if role in ["employee", "hod", "hod_coordinator"] else None,
                    employee_id=user_id,
                )

                return JsonResponse({"success": True})

        except IntegrityError as e:
            if 'unique_user_type_per_user' in str(e):
                return JsonResponse({
                    "success": False, 
                    "error": f"User already has a {role} profile"
                })
            return JsonResponse({
                "success": False, 
                "error": str(e)
            })
        except Exception as e:
            return JsonResponse({
                "success": False, 
                "error": str(e)
            })

    departments = [dept[0] for dept in Profile.DEPARTMENT_CHOICES]
    return render(request, "dashboard/register.html", {"departments": departments})


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user_type = request.POST.get('user_type')
        
        print(f"Login attempt - Username: {username}, User Type: {user_type}")  # Debug log

        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            print(f"User authenticated successfully: {user.username}")  # Debug log
            # Check if user is admin
            if username.lower() == 'admin':
                if user.profiles.filter(user_type='admin').exists():
                    login(request, user)
                    request.session['active_profile_type'] = 'admin'
                    return redirect('admin_dashboard')
                else:
                    messages.error(request, 'Invalid admin credentials.')
            # Regular user login
            elif user.profiles.filter(user_type=user_type).exists():
                login(request, user)
                request.session['active_profile_type'] = user_type
                redirects = {
                    'employee': 'employee_dashboard',
                    'hod': 'hod_dashboard',
                    'coordinator': 'coordinator_dashboard',
                    'finance': 'finance_dashboard',
                    'hod_coordinator': 'hod_coordinator_dashboard'  
                }
                return redirect(redirects.get(user_type, 'login'))
            else:
                messages.error(request, f'You do not have {user_type} access.')
        else:
            messages.error(request, 'Invalid username or password.')
            
    return render(request, 'dashboard/login.html')

# Logout view
def logout_view(request):
    logout(request)
    return redirect("login")


# Dashboard view
@login_required
def dashboard_view(request):
    if hasattr(request.user, "profile"):
        if request.active_profile.is_coordinator:
            return redirect("coordinator_dashboard")
        else:
            return redirect("employee_dashboard")
    return render(
        request, "dashboard/view_kaizen.html"
    )  # Changed from view_kaizen_sheets.html


@login_required
def update_profile(request):
    if request.method == 'POST':
        try:
            user = request.user
            username = request.POST.get('username')
            password = request.POST.get('password')
            department = request.POST.get('department')

            # Check if username is taken by another user
            if User.objects.exclude(id=user.id).filter(username=username).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'Username already taken'
                })

            user.username = username
            
            if password:
                user.set_password(password)
                # Get the active profile instead of trying to access profile directly
                active_profile = user.profiles.filter(id=request.session.get('active_profile_id')).first()
                if active_profile:
                    active_profile.last_password_change = timezone.now()
                    active_profile.save()

            if request.active_profile.user_type in ['employee', 'hod']:
                if request.active_profile.department != department:
                    old_department = request.active_profile.department
                    request.active_profile.department = department
                    request.active_profile.last_department_change = timezone.now()
                    request.active_profile.save()
                    
                    print(f"User {username} changed department from {old_department} to {department}")

            user.save()

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })

# views.py
@login_required
@profile_required('employee')
def employee_dashboard(request):
    if not request.active_profile.is_employee:
        return HttpResponseForbidden()
    

    if request.method == "POST":
        title = request.POST.get("title")
        instance = KaizenSheet.objects.filter(
            title=title, employee=request.user, is_temporary=False
        ).first()

        if instance and instance.approval_status != "pending":
            messages.error(request, "Cannot edit an approved kaizen sheet")
            return redirect("employee_dashboard")

        form = KaizenSheetForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            kaizen = form.save(commit=False)
            kaizen.employee = request.user
            kaizen.is_temporary = False
            kaizen.save()

            horizontal_deployment = (
                request.POST.get("id_horizontal_deployment") == "on"
            )
            if horizontal_deployment:
                selected_departments = request.POST.getlist(
                    "horizontal_departments"
                )
                # Clear existing deployments
                HorizontalDeployment.objects.filter(
                    kaizen_sheet=kaizen
                ).delete()

                # Create new deployments
                for dept in selected_departments:
                    print(f"Creating deployment for department: {dept}")
                    HorizontalDeployment.objects.create(
                        kaizen_sheet=kaizen, department=dept
                    )
            # Handle file uploads
            for field in [
                "before_improvement_image",
                "after_improvement_image",
                "standardization_file",
                "cost_calculation",
            ]:
                if field in request.FILES:
                    setattr(kaizen, field, request.FILES[field])

            # Handle impact fields
            impacts = [
                "safety",
                "quality",
                "productivity",
                "delivery",
                "cost",
                "morale",
                "environment",
            ]
            for impact in impacts:
                impact_checked = request.POST.get(f"impacts_{impact}") == "true"
                setattr(kaizen, f"impacts_{impact}", impact_checked)

                if impact_checked:
                    for field_type in [
                        "benefits_description",
                        "uom",
                        "before_implementation",
                        "after_implementation",
                    ]:
                        field_name = f"{impact}_{field_type}"
                        if field_name in request.POST:
                            setattr(
                                kaizen, field_name, request.POST.get(field_name)
                            )

            kaizen.save()
            messages.success(request, "Kaizen sheet saved successfully")
            return redirect("employee_dashboard")
    else:
        form = KaizenSheetForm()

    current_department = request.active_profile.department

    # Get all kaizen sheets for the user
    kaizen_sheets = KaizenSheet.objects.filter(
        employee=request.user,
        is_temporary=False
    ).order_by("-created_at")

    # Get handwritten sheets
    handwritten_sheets = KaizenSheet.objects.filter(
        employee=request.user,
        is_handwritten=True
    ).order_by('-created_at')

    # Get available departments for horizontal deployment
    departments = (
        Profile.objects.filter(user_type="hod")
        .exclude(department=request.active_profile.department)
        .values_list("department", flat=True)
        .distinct()
    )

    department_choices = [dept[0] for dept in Profile.DEPARTMENT_CHOICES]


    context = {
        "form": form,
        "kaizen_sheets": kaizen_sheets,
        "departments": departments,
        "kaizen_list": kaizen_sheets,
        "pending_sheets": kaizen_sheets.filter(approval_status="pending"),
        'handwritten_sheets': handwritten_sheets,
        "department_choices": department_choices,  # For profile editing
        'area_grouping_choices': KaizenSheet.AREA_GROUPING_CHOICES,
        "hod_approved_sheets": kaizen_sheets.filter(
            approval_status="hod_approved"
        ),
        "finance_pending_sheets": kaizen_sheets.filter(
            approval_status="finance_pending"
        ),
        "finance_approved_sheets": kaizen_sheets.filter(
            approval_status="finance_approved"
        ),
        "coordinator_approved_sheets": kaizen_sheets.filter(
            approval_status="coordinator_approved"
        ),
        "impacts": [
            "safety",
            "quality",
            "productivity",
            "delivery",
            "cost",
            "morale",
            "environment",
        ],
    }

    return render(request, "dashboard/employee_dashboard.html", context)


@login_required
def approve_kaizen(request, kaizen_id):
    if not request.active_profile.is_hod:
        return JsonResponse({"success": False, "error": "Unauthorized"})

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid method"})

    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)

    try:
        # Update HOD approval fields
        sheet.hod_approved = True
        sheet.hod_approved_by = request.user
        sheet.hod_approved_at = timezone.now()

        # Calculate cost difference for next approval step
        cost_diff = sheet.get_cost_difference()

        # Determine next approval step based on cost
        if cost_diff > 100000:
            sheet.approval_status = "finance_pending"
        elif cost_diff > 45000:
            sheet.approval_status = "coordinator_pending"
        else:
            sheet.approval_status = "completed"

        sheet.save()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

@login_required
def view_sheet(request, sheet_id):
    if not request.active_profile.is_hod:
        return HttpResponseForbidden()

    sheet = get_object_or_404(
        KaizenSheet, id=sheet_id, horizontal_departments=request.active_profile.department
    )

    return render(request, "dashboard/view_kaizen.html", {"sheet": sheet})

@login_required
def fetch_kaizen_sheet(request, kaizen_id):
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        try:
            kaizen = get_object_or_404(KaizenSheet, id=kaizen_id)
            
            # Get deployed departments
            deployed_depts = HorizontalDeployment.objects.filter(
                kaizen_sheet=kaizen
            ).values_list("department", flat=True)

            # Basic data
            data = {
                "title": kaizen.title,
                "area_implemented": kaizen.area_implemented,
                "area_grouping": kaizen.area_grouping,
                "start_date": kaizen.start_date.strftime("%Y-%m-%d") if kaizen.start_date else "",
                "end_date": kaizen.end_date.strftime("%Y-%m-%d") if kaizen.end_date else "",
                "problem": kaizen.problem,
                "idea_solved": kaizen.idea_solved,
                "standardization": kaizen.standardization,
                "deployment": kaizen.deployment,
                "team_member2_id": kaizen.team_member2_id,
                "team_member2": kaizen.team_member2,
                "horizontal_departments": list(deployed_depts),
                "before_improvement_text": kaizen.before_improvement_text,
                "after_improvement_text": kaizen.after_improvement_text,
            }

            # Add file URLs if they exist
            if kaizen.before_improvement_image:
                data["before_improvement_image"] = kaizen.before_improvement_image.url
            if kaizen.after_improvement_image:
                data["after_improvement_image"] = kaizen.after_improvement_image.url
            if kaizen.standardization_file:
                data["standardization_file"] = kaizen.standardization_file.url
            if kaizen.cost_calculation:
                data["cost_calculation"] = kaizen.cost_calculation.url

            # Add impact fields
            impacts = ['safety', 'quality', 'productivity', 'delivery', 'cost', 'morale', 'environment']
            for impact in impacts:
                # Add impact checkbox state
                data[f'impacts_{impact}'] = getattr(kaizen, f'impacts_{impact}')
                
                # Add impact fields
                for field in ['benefits_description', 'uom', 'before_implementation', 'after_implementation']:
                    field_name = f'{impact}_{field}'
                    data[field_name] = getattr(kaizen, field_name, '')

            return JsonResponse(data)
            
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
            
    return HttpResponseBadRequest("Invalid request")


# views.py
@login_required
def create_temp_kaizen(request):
    if request.method == "POST":
        form = KaizenSheetForm(request.POST, request.FILES)
        if form.is_valid():
            title = form.cleaned_data["title"]

            # Check for existing non-temporary sheet
            existing_sheet = KaizenSheet.objects.filter(
                title=title, employee=request.user, is_temporary=False
            ).first()

            if existing_sheet:
                # Return existing sheet ID for download
                return JsonResponse({"kaizen_id": existing_sheet.id})
            else:
                if not title:
                    return JsonResponse(
                        {
                            "error": "Please fill required fields before downloading"
                        },
                        status=400,
                    )

                # Create temporary sheet for download
                kaizen = form.save(commit=False)
                kaizen.employee = request.user
                kaizen.is_temporary = True
                kaizen.save()

                return JsonResponse({"kaizen_id": kaizen.id})

    return JsonResponse({"error": "Invalid form data"}, status=400)


# views.py
def check_impact_data(kaizen, impact):
    return any(
        [
            getattr(kaizen, f"{impact}_benefits_description", ""),
            getattr(kaizen, f"{impact}_uom", ""),
            getattr(kaizen, f"{impact}_before_implementation", ""),
            getattr(kaizen, f"{impact}_after_implementation", ""),
        ]
    )


@login_required
def download_kaizen_sheet(request, kaizen_id):
    try:
        kaizen = get_object_or_404(
            KaizenSheet, id=kaizen_id, employee=request.user
        )

        # Load template
        template_path = "kaizen_app/templates/excel/kaizen_format.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active

        def write_to_merged_cell(
            sheet, cell_address, value, is_yes_no=False, font_style=None
        ):
            for merged_range in sheet.merged_cells.ranges:
                if cell_address in merged_range:
                    sheet.unmerge_cells(str(merged_range))
                    cell = sheet[cell_address]
                    cell.value = value
                    if is_yes_no:
                        cell.font = Font(name="Arial", size=11)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    elif font_style:
                        cell.font = font_style
                    sheet.merge_cells(str(merged_range))
                    return

            # If not merged, just set the value
            cell = sheet[cell_address]
            cell.value = value
            if is_yes_no:
                cell.font = Font(name="Arial", size=11)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            elif font_style:
                cell.font = font_style

        # Basic info
        write_to_merged_cell(ws, "C2", kaizen.title)
        write_to_merged_cell(ws, "T3", kaizen.serial_key)
        write_to_merged_cell(ws, "A5", kaizen.area_implemented)
        write_to_merged_cell(ws, "E5", kaizen.start_date.strftime("%Y-%m-%d"))
        write_to_merged_cell(ws, "H5", kaizen.end_date.strftime("%Y-%m-%d"))
        employee_info = f"{kaizen.employee.username} - {kaizen.employee.profiles.first().employee_id}"
        write_to_merged_cell(ws, "S4", employee_info)

        # Problem section
        write_to_merged_cell(ws, "A7", kaizen.problem)

        if kaizen.hod_approved and kaizen.hod_approved_by:
            write_to_merged_cell(
                ws,
                "K5",
                kaizen.hod_approved_by.get_full_name() or kaizen.hod_approved_by.username,
                font_style=Font(name="Arial", size=11, bold=False)
            )   
        
        # Before improvement section
        if kaizen.before_improvement_image:
            try:
                image_path = os.path.join(
                    settings.MEDIA_ROOT, str(kaizen.before_improvement_image)
                )
                if os.path.exists(image_path):
                    img = XLImage(image_path)
                    img.width = 625
                    img.height = 450
                    ws.add_image(img, "G7")
            except Exception as e:
                print(f"Error adding before image: {str(e)}")
        write_to_merged_cell(
            ws,
            "G16",
            kaizen.before_improvement_text,
            font_style=Font(name="Arial", size=11, bold=False),
        )

        # Ideas section
        write_to_merged_cell(ws, "A21", kaizen.idea_solved)

        # After improvement section
        if kaizen.after_improvement_image:
            try:
                image_path = os.path.join(
                    settings.MEDIA_ROOT, str(kaizen.after_improvement_image)
                )
                if os.path.exists(image_path):
                    img = XLImage(image_path)
                    img.width = 625
                    img.height = 450
                    ws.add_image(img, "G21")
            except Exception as e:
                print(f"Error adding after image: {str(e)}")
        write_to_merged_cell(
            ws,
            "G32",
            kaizen.after_improvement_text,
            font_style=Font(name="Arial", size=11, bold=False),
        )

        # Impact matrix with correct cell mappings
        impacts = [
            ("safety", 10),  # P10-V10
            ("quality", 12),  # P12-V12
            ("productivity", 14),  # P14-V14
            ("delivery", 16),  # P16-V16
            ("cost", 18),  # P18-V18
            ("morale", 20),  # P20-V20
            ("environment", 22),  # P22-V22
        ]

        for impact_name, row_num in impacts:
            # Check if impact has data
            is_checked = check_impact_data(kaizen, impact_name)
            print(f"Impact {impact_name}: {is_checked}")  # Debug log

            # Write Yes/No with special handling
            write_to_merged_cell(
                ws, f"P{row_num}", "Yes" if is_checked else "No", is_yes_no=True
            )

            # Write impact data fields
            write_to_merged_cell(
                ws,
                f"Q{row_num}",
                getattr(kaizen, f"{impact_name}_benefits_description", ""),
            )
            write_to_merged_cell(
                ws,
                f"S{row_num}",
                getattr(kaizen, f"{impact_name}_uom", "") or "",
            )
            write_to_merged_cell(
                ws,
                f"T{row_num}",
                getattr(kaizen, f"{impact_name}_before_implementation", ""),
            )
            write_to_merged_cell(
                ws,
                f"V{row_num}",
                getattr(kaizen, f"{impact_name}_after_implementation", ""),
            )

        # Standardization and deployment
        write_to_merged_cell(ws, "N25", kaizen.standardization)
        write_to_merged_cell(ws, "N32", kaizen.deployment)

        # Set response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=Kaizen_Sheet_{kaizen.serial_key}.xlsx"
        )
        wb.save(response)
        return response

    except Exception as e:
        print(f"Error in download_kaizen_sheet: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def edit_kaizen_sheet(request, sheet_id):
    kaizen_sheet = get_object_or_404(
        KaizenSheet, id=sheet_id, employee=request.user
    )

    if kaizen_sheet.is_approved:
        return redirect("dashboard")

    if request.method == "POST":
        form = KaizenSheetForm(
            request.POST, request.FILES, instance=kaizen_sheet
        )
        if form.is_valid():
            form.save()
            return redirect("dashboard")
    else:
        form = KaizenSheetForm(instance=kaizen_sheet)

    return render(
        request,
        "dashboard/edit_kaizen_sheet.html",
        {"form": form, "sheet_id": sheet_id},
    )

@login_required 
def update_handwritten_sheet(request):
    if request.method == 'POST':
        try:
            sheet_id = request.POST.get('sheet_id')
            sheet = get_object_or_404(KaizenSheet, id=sheet_id, employee=request.user)
            
            # Check if sheet is already approved
            if sheet.approval_status not in ['pending', None]:
                return JsonResponse({
                    'success': False, 
                    'error': 'Cannot edit approved sheets'
                })

            # Update title
            new_title = request.POST.get('handwritten_title')
            if new_title:
                sheet.title = new_title

            # Update sheet file if provided
            new_sheet = request.FILES.get('handwritten_sheet')
            if new_sheet:
                # Delete old file if exists
                if sheet.handwritten_sheet:
                    sheet.handwritten_sheet.delete()
                sheet.handwritten_sheet = new_sheet

            sheet.save()
            return JsonResponse({'success': True})

        except KaizenSheet.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Sheet not found'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })

@login_required
def delete_account(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                user = request.user
                # Get active profile before deletion
                active_profile = request.active_profile
                
                if not active_profile:
                    return JsonResponse({
                        'success': False,
                        'error': 'No active profile found'
                    })

                # Store profile info for kaizen sheets
                profile_info = {
                    'username': 'Account Deleted',
                    'employee_id': 'Deleted',
                    'department': active_profile.department if active_profile else None
                }

                # Anonymize kaizen sheets only if employee profile is being deleted
                if active_profile.user_type == 'employee':
                    KaizenSheet.objects.filter(employee=user).update(
                        employee_info=profile_info
                    )

                # Delete only the active profile
                active_profile.delete()

                # If this was user's last profile, anonymize and deactivate user
                if not user.profiles.exists():
                    user.username = f"deleted_user_{user.id}"
                    user.email = f"deleted_{user.id}@deleted.com"
                    user.is_active = False
                    user.save()
                    logout(request)
                    return JsonResponse({
                        'success': True,
                        'message': 'Account deleted successfully',
                        'redirect': True
                    })
                else:
                    # If user has other profiles, just redirect to login to switch profiles
                    logout(request)
                    return JsonResponse({
                        'success': True,
                        'message': f'{active_profile.user_type.title()} profile deleted successfully',
                        'redirect': True
                    })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })

@login_required
def save_kaizen_sheet(request):
    if request.method == "POST":
        kaizen_id = request.POST.get('kaizen_id')
        instance = None
        
        if kaizen_id:
            instance = get_object_or_404(KaizenSheet, id=kaizen_id, employee=request.user)
            
        form = KaizenSheetForm(request.POST, request.FILES, instance=instance)
        
        if form.is_valid():
            try:
                kaizen = form.save(commit=False)
                
                # Check CIP number uniqueness if it's changed or new
                if instance:
                    old_cip = instance.serial_key
                    new_cip = form.cleaned_data.get('serial_key')
                    if old_cip != new_cip:
                        if KaizenSheet.objects.exclude(id=instance.id).filter(serial_key=new_cip).exists():
                            return JsonResponse({
                                'success': False, 
                                'error': 'This CIP number already exists. Please use a different one.'
                            })
                else:
                    if KaizenSheet.objects.filter(serial_key=form.cleaned_data.get('serial_key')).exists():
                        return JsonResponse({
                            'success': False, 
                            'error': 'This CIP number already exists. Please use a different one.'
                        })

                kaizen.employee = request.user
                
                # Handle impact fields
                impact_types = ['safety', 'quality', 'productivity', 'delivery', 'cost', 'morale', 'environment']
                
                for impact in impact_types:
                    # Save impact checkbox state
                    impact_checked = request.POST.get(f'id_impacts_{impact}') == 'on'
                    setattr(kaizen, f'impacts_{impact}', impact_checked)
                    
                    # Save impact details if checked
                    if impact_checked:
                        fields = ['benefits_description', 'uom', 'before_implementation', 'after_implementation']
                        for field in fields:
                            field_name = f'{impact}_{field}'
                            field_value = request.POST.get(f'id_{impact}_{field}')
                            setattr(kaizen, field_name, field_value)
                
                kaizen.save()
                
                # Handle file uploads
                file_fields = [
                    'standardization_file', 
                    'cost_calculation',
                    'before_improvement_image', 
                    'after_improvement_image'
                ]
                
                for field in file_fields:
                    if field in request.FILES:
                        # Delete old file if exists
                        old_file = getattr(kaizen, field, None)
                        if old_file:
                            old_file.delete()
                        # Save new file    
                        setattr(kaizen, field, request.FILES[field])
                
                # Final save after file handling
                kaizen.save()

                # If this is a handwritten sheet, handle additional fields
                if kaizen.is_handwritten:
                    if 'handwritten_sheet' in request.FILES:
                        old_sheet = kaizen.handwritten_sheet
                        if old_sheet:
                            old_sheet.delete()
                        kaizen.handwritten_sheet = request.FILES['handwritten_sheet']
                        kaizen.save()

                return JsonResponse({
                    'success': True,
                    'kaizen_id': kaizen.id,
                    'serial_key': kaizen.serial_key
                })
                
            except IntegrityError as e:
                if 'serial_key' in str(e):
                    return JsonResponse({
                        'success': False, 
                        'error': 'This CIP number already exists. Please use a different one.'
                    })
                return JsonResponse({
                    'success': False, 
                    'error': f'Database error: {str(e)}'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False, 
                    'error': f'Error saving kaizen sheet: {str(e)}'
                })
        else:
            return JsonResponse({
                'success': False, 
                'error': form.errors
            })
            
    return JsonResponse({
        'success': False, 
        'error': 'Invalid request method'
    })

# views.py
@login_required
def approve_kaizen(request, kaizen_id):
    if not request.active_profile.is_hod:
        return JsonResponse({"success": False, "error": "Unauthorized"})

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid method"})

    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)

    try:
        sheet.hod_approved = True
        sheet.hod_approved_by = request.user
        sheet.hod_approved_at = timezone.now()

        cost_diff = sheet.get_cost_difference()

        # Update status based on cost difference
        if cost_diff <= 45000:
            sheet.approval_status = "completed"
        elif cost_diff <= 100000:
            sheet.approval_status = "coordinator_pending"
        else:
            sheet.approval_status = "finance_pending"

        sheet.save()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def get_department_stats(dept_name):
    """Helper function to get department statistics"""
    if not dept_name:
        return None

    # Update HOD query
    hod = User.objects.filter(
        profiles__department=dept_name, 
        profiles__user_type="hod"
    ).first()

    # Update employee count query
    employee_count = Profile.objects.filter(
        department=dept_name, 
        user_type="employee"
    ).count()

    try:
        coordinator_entry = KaizenCoordinator.objects.get_or_create(
            department=dept_name, defaults={"coordinator_name": ""}
        )[0]
        coordinator_name = coordinator_entry.coordinator_name or ""
    except Exception:
        coordinator_name = ""

    # Update completed sheets query
    completed_sheets = KaizenSheet.objects.filter(
        employee__profiles__department=dept_name,
        approval_status="completed",
        hod_approved_at__isnull=False,
    ).count()

    monthly_target = employee_count
    achievement_rate = (
        (completed_sheets / monthly_target * 100) if monthly_target > 0 else 0
    )

    return {
        "name": dept_name,
        "hod_name": hod.username if hod else "No HOD Assigned",
        "kaizen_coordinator": coordinator_name,
        "employee_count": employee_count,
        "monthly_target": monthly_target,
        "completed": completed_sheets,
        "achievement_rate": round(achievement_rate, 2),
    }



@login_required
@profile_required('coordinator')
def coordinator_dashboard(request):
    if not request.active_profile.is_coordinator:
        return HttpResponseForbidden()

    try:
        # Get current timestamp and earliest kaizen
        today = timezone.localtime(timezone.now())
        current_year = today.year
        current_month = today.month
        
        earliest_kaizen = KaizenSheet.objects.order_by('created_at').first()
        start_year = earliest_kaizen.created_at.year if earliest_kaizen else current_year

        # Academic year calculations
        academic_start_year = current_year if current_month >= 4 else current_year - 1
        academic_years_data = []

        for year in range(start_year, current_year + 1):
            start_date = timezone.make_aware(timezone.datetime(year, 4, 1))
            end_date = timezone.make_aware(timezone.datetime(year + 1, 3, 31))

            if end_date > today:
                end_date = today

            completed_sheets = KaizenSheet.objects.filter(
                hod_approved_at__range=(start_date, end_date),
                approval_status="completed",
            )

            months_diff = (
                (end_date.year - start_date.year) * 12
                + end_date.month
                - start_date.month
                + 1
            )

            monthly_average = round(completed_sheets.count() / months_diff, 2) if months_diff > 0 else 0
            academic_years_data.append({
                "year": f"AY {year}-{year + 1}",
                "average": monthly_average
            })

        # Monthly calculations
        monthly_submissions = [0] * 12
        monthly_completions = [0] * 12
        monthly_averages = [0] * 12
        cumulative_completed = 0

        total_employees = User.objects.filter(
            profiles__user_type='employee'
        ).count()
        
        # Calculate monthly statistics
        for month in range(1, 13):
            month_start = timezone.make_aware(timezone.datetime(current_year, month, 1))
            month_end = timezone.make_aware(
                timezone.datetime(current_year, month + 1, 1) if month < 12 
                else timezone.datetime(current_year + 1, 1, 1)
            ) - timezone.timedelta(days=1)

            if month_start <= today:
                submissions = KaizenSheet.objects.filter(
                    created_at__range=(month_start, month_end)
                ).count()

                completions = KaizenSheet.objects.filter(
                    hod_approved_at__range=(month_start, month_end),
                    approval_status="completed",
                ).count()

                monthly_submissions[month - 1] = submissions
                monthly_completions[month - 1] = completions
                
                cumulative_completed += completions
                monthly_averages[month - 1] = round(cumulative_completed / month, 2)

        # Department statistics
        departments = []
        department_names = Profile.objects.values_list(
            "department", flat=True
        ).distinct().exclude(department__isnull=True)

        for dept_name in department_names:
            dept_data = get_department_stats(dept_name)
            if dept_data:
                departments.append(dept_data)

        # Employee statistics
        employee_stats = []
        for employee in User.objects.filter(profiles__user_type='employee'):
            all_sheets = KaizenSheet.objects.filter(employee=employee)
            department = employee.profiles.filter(user_type='employee').first().department
            completed = all_sheets.filter(approval_status='completed').count()
            pending = all_sheets.exclude(approval_status='completed').count()

            employee_stats.append({
                'username': employee.get_full_name() or employee.username,
                'department': department,
                'total_submissions': all_sheets.count(),
                'approved': completed,
                'pending': pending
            })
        # Debug prints for yearly performance data
        print("Monthly Completions:", monthly_completions)
        print("Current Month:", current_month)
        print("Current Year:", current_year)
        context = {
            "current_year": current_year,
            "year_range": range(current_year, start_year - 1, -1),
            "months": json.dumps([
                "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
            ]),
            "monthly_submissions": json.dumps(monthly_submissions, cls=DjangoJSONEncoder),
            "monthly_completions": json.dumps(monthly_completions, cls=DjangoJSONEncoder),
            "monthly_averages": json.dumps(monthly_averages, cls=DjangoJSONEncoder),
            "academic_years": json.dumps([d["year"] for d in academic_years_data]),
            "academic_year_counts": json.dumps([d["average"] for d in academic_years_data], cls=DjangoJSONEncoder),
            "current_academic_year": f"{academic_start_year}-{academic_start_year + 1}",
            "current_month": current_month,
            "total_employees": total_employees,
            "departments": departments,
            "total_kaizens": KaizenSheet.objects.count(),
            "completed_count": KaizenSheet.objects.filter(approval_status="completed").count(),
            "hod_pending_count": KaizenSheet.objects.filter(approval_status="pending").count(),
            "coordinator_pending_count": KaizenSheet.objects.filter(approval_status="coordinator_pending").count(),
            "finance_pending_count": KaizenSheet.objects.filter(approval_status="finance_pending").count(),
            "pending_approvals": KaizenSheet.objects.exclude(approval_status="completed").count(),
            "coordinator_approved": KaizenSheet.objects.filter(approval_status="completed").count(),
            "employee_stats": employee_stats,
            "department_data": json.dumps(departments, cls=DjangoJSONEncoder),
        }

        return render(request, "dashboard/coordinator_dashboard.html", context)

    except Exception as e:
        messages.error(request, f"Error loading dashboard: {str(e)}")
        return redirect('login')


@login_required
@require_POST
def save_kaizen_coordinators(request):
    if not request.active_profile.is_coordinator:
        return JsonResponse({"success": False, "error": "Access denied."})

    try:
        data = json.loads(request.body)
        for department, coordinator_name in data.items():
            KaizenCoordinator.objects.update_or_create(
                department=department,
                defaults={"coordinator_name": coordinator_name},
            )
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required
def get_yearly_data(request, year):
    year_type = request.GET.get('year_type', 'calendar')
    try:
        year = int(year)
        
        if year_type == 'financial':
            # Financial year is April to March
            start_date = timezone.datetime(year, 4, 1)  # April 1st
            end_date = timezone.datetime(year + 1, 3, 31, 23, 59, 59)  # March 31st next year
        else:
            # Calendar year
            start_date = timezone.datetime(year, 1, 1)
            end_date = timezone.datetime(year, 12, 31, 23, 59, 59)

        start_date = timezone.make_aware(start_date)
        end_date = timezone.make_aware(end_date)

        # Query all kaizen sheets within date range
        kaizen_sheets = KaizenSheet.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        )

        # Initialize arrays for monthly data
        submissions = [0] * 12
        completions = [0] * 12

        # Process each kaizen sheet
        for sheet in kaizen_sheets:
            # For financial year, shift month index (April = 0, March = 11)
            if year_type == 'financial':
                month_index = (sheet.created_at.month - 4) % 12
            else:
                month_index = sheet.created_at.month - 1
                
            submissions[month_index] += 1
            if sheet.approval_status == 'completed':
                completions[month_index] += 1

        return JsonResponse({
            'success': True,
            'submissions': submissions,
            'completions': completions,
            'year_type': year_type,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d')
        })

    except Exception as e:
        print(f"Error in get_yearly_data: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
    
@login_required
def check_historical_data(request, year):
    year_type = request.GET.get('year_type', 'calendar')
    try:
        year = int(year)
        historical_data = {'has_historical': False}
        
        if year_type == 'financial':
            # Check last 4 financial years
            for prev_year in range(year-4, year):
                # Get data from April of prev_year to March of next year
                start_date = datetime(prev_year, 4, 1)
                end_date = datetime(prev_year + 1, 3, 31)
                
                completions = KaizenSheet.objects.filter(
                    created_at__gte=start_date,
                    created_at__lte=end_date,
                    approval_status='completed'
                )
                
                if completions.exists():
                    historical_data['has_historical'] = True
                    # Calculate average completions per month
                    total_completions = completions.count()
                    months = min(12, (end_date - start_date).days // 30)
                    historical_data[f'fy_{prev_year}'] = total_completions / months
        else:
            # Check last 4 calendar years
            for prev_year in range(year-4, year):
                completions = KaizenSheet.objects.filter(
                    created_at__year=prev_year,
                    approval_status='completed'
                )
                
                if completions.exists():
                    historical_data['has_historical'] = True
                    total_completions = completions.count()
                    historical_data[f'year_{prev_year}'] = total_completions / 12

        return JsonResponse(historical_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

def get_all_financial_years(request):
    try:
        yearly_data = {}
        current_date = datetime.now()
        
        # Get all completed kaizen sheets
        kaizen_sheets = KaizenSheet.objects.filter(
            approval_status='completed'
        ).order_by('created_at')

        # Group data by financial year
        for sheet in kaizen_sheets:
            sheet_date = sheet.created_at
            # Financial year starts from April
            financial_year = sheet_date.year if sheet_date.month >= 4 else sheet_date.year - 1
            
            if financial_year not in yearly_data:
                yearly_data[financial_year] = {
                    'total': 0,
                    'months': set(),
                    'monthly_counts': [0] * 12  # Track monthly counts
                }
            
            # Map the month to financial year order (Apr-Mar)
            month_index = (sheet_date.month - 4) % 12
            yearly_data[financial_year]['monthly_counts'][month_index] += 1
            yearly_data[financial_year]['total'] += 1
            yearly_data[financial_year]['months'].add(sheet_date.month)

        # Calculate averages
        final_data = {}
        for year, data in yearly_data.items():
            month_count = len(data['months'])
            if month_count > 0:
                final_data[year] = data['total'] / month_count

        return JsonResponse({
            'success': True,
            'yearly_data': final_data,
            'raw_monthly_data': {
                year: data['monthly_counts'] 
                for year, data in yearly_data.items()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def check_cip_number(request, cip_number, current_sheet_id):
    """Check if a CIP number exists, excluding the current sheet"""
    try:
        # Check if CIP exists in any other sheet
        exists = KaizenSheet.objects.exclude(
            id=current_sheet_id
        ).filter(
            serial_key=cip_number,
            is_temporary=False  # Exclude temporary sheets
        ).exists()
        
        return JsonResponse({'exists': exists})
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)
    
@login_required
def get_department_data(request, year, month):
    if not request.active_profile.is_coordinator:
        return JsonResponse({"error": "Unauthorized"}, status=403)

    try:
        # Convert month and year to date range
        start_date = timezone.make_aware(datetime(int(year), int(month), 1))
        if int(month) == 12:
            end_date = timezone.make_aware(datetime(int(year) + 1, 1, 1))
        else:
            end_date = timezone.make_aware(datetime(int(year), int(month) + 1, 1))

        print(f"Date range: {start_date} to {end_date}")  # Debug log

        departments = []
        department_names = Profile.objects.values_list(
            "department", flat=True
        ).distinct().exclude(department__isnull=True)

        for dept_name in department_names:
            if not dept_name:
                continue

            # Get monthly completions - add debug logs
            completed_this_month = KaizenSheet.objects.filter(
                employee__profiles__department=dept_name,
                approval_status="completed",
                hod_approved_at__range=(start_date, end_date)
            )
            
            # Debug logs
            print(f"\nDepartment: {dept_name}")
            print(f"Completions query: {completed_this_month.query}")
            print(f"Completed count: {completed_this_month.count()}")
            
            # Get total completions for debugging
            total_completed = KaizenSheet.objects.filter(
                employee__profiles__department=dept_name,
                approval_status="completed"
            ).count()
            
            print(f"Total completed: {total_completed}")

            # Get employee count for target
            employee_count = Profile.objects.filter(
                department=dept_name,
                user_type="employee"
            ).count()
            
            print(f"Employee count: {employee_count}")
            print(f"Monthly target: {employee_count}")

            # Calculate metrics
            monthly_target = employee_count
            completed_count = completed_this_month.count()
            achievement_rate = (completed_count / monthly_target * 100) if monthly_target > 0 else 0
            
            print(f"Achievement rate: {achievement_rate}%")

            departments.append({
                "name": dept_name,
                "hod_name": User.objects.filter(
                    profiles__department=dept_name,
                    profiles__user_type="hod"
                ).first().username if User.objects.filter(
                    profiles__department=dept_name,
                    profiles__user_type="hod"
                ).exists() else "No HOD Assigned",
                "kaizen_coordinator": KaizenCoordinator.objects.filter(
                    department=dept_name
                ).first().coordinator_name if KaizenCoordinator.objects.filter(
                    department=dept_name
                ).exists() else "",
                "monthly_target": monthly_target,
                "completed": completed_count,
                "total_completed": total_completed,
                "achievement_rate": achievement_rate
            })

        return JsonResponse({"success": True, "departments": departments})

    except Exception as e:
        print(f"Error in get_department_data: {str(e)}")  # Debug log
        return JsonResponse({"success": False, "error": str(e)})

@login_required
def get_excel_template(request, template_name):
    try:
        file_path = os.path.join(
            settings.BASE_DIR, "kaizen_app", "templates", "excel", template_name
        )
        if os.path.exists(file_path):
            with open(file_path, "rb") as template:
                response = HttpResponse(
                    template.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = (
                    f'attachment; filename="{template_name}"'
                )
                return response
        return HttpResponse("Template not found", status=404)
    except Exception as e:
        return HttpResponse(str(e), status=500)


@login_required
@profile_required('coordinator')
def cip_register_view(request):
    try:
        print(f"Active Profile: {request.active_profile}")
        if not request.active_profile or not request.active_profile.is_coordinator:
            messages.error(request, "Access denied. Coordinator privileges required.")
            return redirect("coordinator_dashboard")

        # Updated query to use prefetch_related for profiles
        kaizen_sheets = KaizenSheet.objects.select_related(
            'employee'
        ).prefetch_related(
            'employee__profiles'
        ).all().order_by("-created_at")

        # Get items per page from request or use default
        items_per_page = request.GET.get('items_per_page', 5)
        try:
            items_per_page = int(items_per_page)
        except ValueError:
            items_per_page = 5

        # Get all kaizen sheets
        kaizen_sheets = KaizenSheet.objects.select_related(
            'employee'
        ).prefetch_related(
            'employee__profiles'
        ).all().order_by("-created_at")

        # Create a dictionary to store department HODs
        department_hods = {}
        for dept in Profile.objects.filter(user_type='hod'):
            department_hods[dept.department] = dept.user.username

        # Attach HOD names to kaizen sheets
        for sheet in kaizen_sheets:
            department = sheet.employee.profiles.first().department
            sheet.project_leader = department_hods.get(department, 'No HOD Assigned')

        # Create paginator
        paginator = Paginator(kaizen_sheets, items_per_page)
        page = request.GET.get('page', 1)

        try:
            kaizen_sheets = paginator.page(page)
        except PageNotAnInteger:
            kaizen_sheets = paginator.page(1)
        except EmptyPage:
            kaizen_sheets = paginator.page(paginator.num_pages)

        impact_types = [
            "safety", "quality", "productivity", 
            "delivery", "cost", "morale", "environment"
        ]

        processed_sheets = []
        for sheet in kaizen_sheets:
            try:
                # Get employee department and HOD 
                employee_profile = sheet.employee.profiles.filter(user_type='employee').first()
                department = employee_profile.department if employee_profile else None
                
                hod = User.objects.filter(
                    profiles__user_type="hod", 
                    profiles__department=department
                ).first()

                # Process sheet data
                sheet_data = {
                    "id": sheet.id,
                    "serial_key": sheet.serial_key,
                    "title": sheet.title,
                    "employee": sheet.employee,
                    "created_at": sheet.created_at,
                    "approval_status": sheet.approval_status,
                    "area_implemented": sheet.area_implemented,
                    "area_grouping": sheet.area_grouping,
                    "start_date": sheet.start_date,
                    "end_date": sheet.end_date,
                    "department": department,
                    "project_leader": hod.username if hod else "No HOD Assigned",
                    "team_member1_id": employee_profile.employee_id if employee_profile else None,
                    "team_member1": sheet.employee.username,
                    "team_member2_id": sheet.team_member2_id,
                    "team_member2": sheet.team_member2,
                    "savings_start_month": sheet.savings_start_month,
                    "estimated_savings": sheet.estimated_savings,
                    "realized_savings": sheet.realized_savings,
                    "standardization_file_url": sheet.standardization_file.url if sheet.standardization_file else None,
                    "cost_calculation_url": sheet.cost_calculation.url if sheet.cost_calculation else None,
                    "before_improvement_image_url": sheet.before_improvement_image.url if sheet.before_improvement_image else None,
                    "after_improvement_image_url": sheet.after_improvement_image.url if sheet.after_improvement_image else None,
                    "is_handwritten": sheet.is_handwritten,
                    "handwritten_sheet_url": sheet.handwritten_sheet.url if sheet.handwritten_sheet else None,
                    "impacts": {
                        impact: getattr(sheet, f'impacts_{impact}', False) 
                        for impact in impact_types
                    },
                    "impact_data": {}
                }

                for impact in impact_types:
                    impact_info = {
                        'benefits_description': getattr(sheet, f'{impact}_benefits_description', ''),
                        'uom': getattr(sheet, f'{impact}_uom', ''),
                        'before_implementation': getattr(sheet, f'{impact}_before_implementation', ''),
                        'after_implementation': getattr(sheet, f'{impact}_after_implementation', '')
                    }
                    sheet_data['impact_data'][impact] = impact_info

                if hasattr(sheet, 'get_cost_difference'):
                    sheet_data['cost_difference'] = sheet.get_cost_difference()
                    sheet_data['cost_details'] = {
                        'benefits_description': sheet.cost_benefits_description,
                        'uom': sheet.cost_uom,
                        'before_implementation': sheet.cost_before_implementation,
                        'after_implementation': sheet.cost_after_implementation
                    }

                processed_sheets.append(sheet_data)

            except Exception as sheet_error:
                print(f"Error processing sheet {sheet.id}: {str(sheet_error)}")
                continue

        context = {
            "kaizen_sheets": kaizen_sheets,
            "impacts": impact_types,
            'items_per_page': items_per_page,
            'area_grouping_choices': KaizenSheet.AREA_GROUPING_CHOICES,
            "months": [
                'January', 'February', 'March', 'April',
                'May', 'June', 'July', 'August', 
                'September', 'October', 'November', 'December'
            ]
        }

        return render(request, "dashboard/cip_register.html", context)

    except Exception as e:
        print(f"CIP Register Error: {str(e)}")
        messages.error(request, f"Error loading CIP register: {str(e)}")
        return redirect("coordinator_dashboard")

@login_required
def update_kaizen_impacts(request, kaizen_id):
    if not request.active_profile.is_coordinator:
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=403)
        
    if request.method != 'POST':
        return JsonResponse({"success": False, "error": "Invalid method"}, status=405)
        
    try:
        data = json.loads(request.body)
        sheet = KaizenSheet.objects.get(id=kaizen_id)
        
        impacts = data.get('impacts', [])
        
        # Update impact fields
        for impact in ['safety', 'quality', 'productivity', 'delivery', 'cost', 'morale', 'environment']:
            if impact in impacts:
                setattr(sheet, f'impacts_{impact}', True)
                # Set minimal data for display
                setattr(sheet, f'{impact}_benefits_description', 'Handwritten submission')
            else:
                setattr(sheet, f'impacts_{impact}', False)
                
        sheet.save()
        return JsonResponse({"success": True})
        
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

# views.py - Update finance approval view
@login_required
def finance_approve_kaizen(request, kaizen_id):
    if not request.active_profile.user_type == "finance":
        return JsonResponse({"success": False, "error": "Unauthorized"})

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid method"})

    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)
    action = request.POST.get("action")

    try:
        if action == "approve":
            sheet.finance_approved = True
            sheet.finance_approved_by = request.user
            sheet.finance_approved_at = timezone.now()
            sheet.approval_status = "coordinator_pending"
        elif action == "reject":
            sheet.finance_approved = False
            sheet.approval_status = "finance_rejected"
        else:
            return JsonResponse({"success": False, "error": "Invalid action"})

        sheet.save()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    
@login_required
@profile_required('hod_coordinator')
def hod_coordinator_dashboard(request):
    if not request.active_profile.is_hod_coordinator:
        return HttpResponseForbidden()

    # Get the active profile's department
    active_profile = request.active_profile
    department = active_profile.department
    
    print(f"Active Profile: {active_profile.user_type}")
    print(f"Department: {department}")

    if not department:
        messages.error(request, "No department assigned to your profile")
        return redirect("login")
    
    # Get all kaizen sheets from department employees, excluding temporary sheets
    dept_kaizens = KaizenSheet.objects.filter(
        employee__profiles__department=department,
        employee__profiles__user_type='employee',
        is_temporary=False
    ).select_related('employee').prefetch_related(
        'employee__profiles'  # Optimize queries
    ).order_by('-created_at')

    # Get horizontal deployments
    horizontal_sheets = (
        KaizenSheet.objects.filter(deployments__department=department)
        .exclude(employee__profiles__department=department)
        .distinct()
        .order_by("-created_at")
    )

    # Calculate department statistics
    dept_stats = {
        "total": dept_kaizens.count(),
        "pending": dept_kaizens.filter(approval_status="pending").count(),
        "completed": dept_kaizens.filter(approval_status="completed").count(),
        "coordinator_pending": dept_kaizens.filter(approval_status="coordinator_pending").count(),
        "finance_pending": dept_kaizens.filter(approval_status="finance_pending").count()
    }

    # Debug information
    print(f"Total Kaizens found: {dept_stats['total']}")
    print(f"Department Query: {dept_kaizens.query}")

    context = {
        "department": department,
        "kaizen_sheets": dept_kaizens,
        "horizontal_sheets": horizontal_sheets,
        "department_stats": dept_stats
    }

    return render(request, "dashboard/hod_coordinator_dashboard.html", context)


@login_required
def coordinator_approve_kaizen(request, kaizen_id):
    if not request.active_profile.is_coordinator:
        return JsonResponse({"success": False, "error": "Unauthorized"})

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid method"})

    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)

    try:
        if sheet.approval_status != "coordinator_pending":
            return JsonResponse(
                {"success": False, "error": "Invalid sheet status"}
            )

        sheet.coordinator_approved = True
        sheet.coordinator_approved_by = request.user
        sheet.coordinator_approved_at = timezone.now()
        sheet.approval_status = "completed"
        sheet.save()

        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required 
def view_kaizen(request, kaizen_id):
    if not (request.active_profile.is_hod or 
            request.active_profile.is_hod_coordinator or 
            request.active_profile.user_type == 'finance'):
        return HttpResponseForbidden()

    try:
        department = request.active_profile.department
        
        # Base queryset with necessary relations
        base_queryset = KaizenSheet.objects.select_related(
            'employee'
        ).prefetch_related(
            'employee__profiles',
            'deployments'
        )

        # HOD and HOD Coordinator can see same sheets from their department
        if request.active_profile.is_hod or request.active_profile.is_hod_coordinator:
            try:
                # Modified query to be more specific
                if request.active_profile.is_hod_coordinator:
                    sheet = base_queryset.filter(
                        id=kaizen_id,
                        employee__profiles__department=department,
                        employee__profiles__user_type='employee'
                    ).first()
                    
                    if not sheet:
                        # Check if it's a horizontal deployment
                        sheet = base_queryset.filter(
                            id=kaizen_id,
                            deployments__department=department
                        ).first()
                        
                    if not sheet:
                        raise KaizenSheet.DoesNotExist()
                        
                    is_horizontal = HorizontalDeployment.objects.filter(
                        kaizen_sheet=sheet,
                        department=department
                    ).exists()
                else:
                    # Original HOD logic
                    sheet = base_queryset.filter(
                        id=kaizen_id
                    ).filter(
                        models.Q(employee__profiles__department=department) |
                        models.Q(deployments__department=department)
                    ).distinct().first()

                    if not sheet:
                        messages.error(request, 'Kaizen sheet not found or unauthorized access')
                        return redirect('hod_dashboard' if request.active_profile.is_hod else 'hod_coordinator_dashboard')

                    # Determine if this is a horizontal deployment
                    is_horizontal = sheet.deployments.filter(department=department).exists()
                    
            except KaizenSheet.DoesNotExist:
                messages.error(request, 'Kaizen sheet not found or unauthorized access')
                return redirect('hod_coordinator_dashboard' if request.active_profile.is_hod_coordinator else 'hod_dashboard')
                
        else:  # Finance user
            sheet = get_object_or_404(base_queryset, id=kaizen_id)
            is_horizontal = False

        # Get source department for reference
        employee_profile = sheet.employee.profiles.filter(user_type='employee').first()
        source_department = employee_profile.department if employee_profile else None

        context = {
            "sheet": sheet,
            "is_horizontal": is_horizontal,
            "source_department": source_department,
            "impacts": [
                "safety", "quality", "productivity", "delivery", 
                "cost", "morale", "environment"
            ],
            "impact_data": {
                impact: {
                    "benefits_description": getattr(sheet, f"{impact}_benefits_description", ""),
                    "uom": getattr(sheet, f"{impact}_uom", ""),
                    "before_implementation": getattr(sheet, f"{impact}_before_implementation", ""),
                    "after_implementation": getattr(sheet, f"{impact}_after_implementation", ""),
                } for impact in ["safety", "quality", "productivity", "delivery", "cost", "morale", "environment"]
            },
            "is_hod": request.active_profile.is_hod,
            "is_hod_coordinator": request.active_profile.is_hod_coordinator,
            "is_coordinator": request.active_profile.is_coordinator,
            "is_finance": request.active_profile.user_type == 'finance',
            "department": department,
            "can_approve": False,  # HOD Coordinator can't approve
            "approval_status": sheet.approval_status
        }

        return render(request, "dashboard/view_kaizen.html", context)

    except Exception as e:
        print(f"Error in view_kaizen: {str(e)}")
        messages.error(request, f"Error accessing kaizen sheet: {str(e)}")
        
        if request.active_profile.user_type == 'finance':
            return redirect('finance_dashboard')
        elif request.active_profile.is_hod_coordinator:
            return redirect('hod_coordinator_dashboard')
        else:
            return redirect('hod_dashboard')
        

@login_required
def reject_kaizen(request, kaizen_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    
    try:
        sheet = KaizenSheet.objects.get(id=kaizen_id)
        user_type = request.active_profile.user_type
        
        if request.active_profile.is_hod:
            sheet.approval_status = 'rejected_by_hod'
        elif user_type == 'coordinator':
            sheet.approval_status = 'rejected_by_coordinator'
        elif user_type == 'finance':
            sheet.approval_status = 'rejected_by_finance'
        else:
            return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
        
        sheet.save()
        return JsonResponse({'success': True})
        
    except KaizenSheet.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Kaizen sheet not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def get_cost_details(request, kaizen_id):
    sheet = get_object_or_404(KaizenSheet, id=kaizen_id)

    data = {
        "benefits_description": sheet.cost_benefits_description,
        "uom": sheet.cost_uom,
        "before_implementation": sheet.cost_before_implementation,
        "after_implementation": sheet.cost_after_implementation,
        "cost_difference": sheet.get_cost_difference(),
        "cost_calculation": (
            sheet.cost_calculation.url if sheet.cost_calculation else None
        ),
    }

    return JsonResponse(data)


@login_required
def update_kaizen(request, kaizen_id):
    if not request.active_profile.is_coordinator:
        return JsonResponse(
            {"success": False, "error": "Unauthorized"}, status=403
        )

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            kaizen = KaizenSheet.objects.get(id=kaizen_id)

            # List of fields that can be updated
            allowed_fields = [
                "title",
                "area_implemented",
                "area_grouping",  # Added area_grouping
                "start_date",
                "end_date",
                "project_leader",
                "team_member1_id",
                "team_member1",
                "team_member2_id",
                "team_member2",
                "savings_start_month",
                "estimated_savings",
                "realized_savings",
            ]

            # Update fields
            for field, value in data.items():
                if field in allowed_fields and hasattr(kaizen, field):
                    # Handle different field types
                    if field in ["estimated_savings", "realized_savings"]:
                        try:
                            value = float(value) if value else 0
                        except ValueError:
                            continue
                    elif field in ["start_date", "end_date"] and value:
                        try:
                            value = datetime.strptime(value, "%Y-%m-%d").date()
                        except ValueError:
                            continue
                    
                    setattr(kaizen, field, value)

            # Special handling for area_grouping
            if 'area_grouping' in data:
                kaizen.area_grouping = data['area_grouping']

            kaizen.save()
            
            response_data = {
                "success": True,
                "area_grouping": kaizen.area_grouping,
                "updated_fields": list(data.keys())
            }
            
            return JsonResponse(response_data)
            
        except KaizenSheet.DoesNotExist:
            return JsonResponse({
                "success": False, 
                "error": "Kaizen sheet not found"
            }, status=404)
        except Exception as e:
            return JsonResponse({
                "success": False, 
                "error": str(e)
            }, status=500)

    return JsonResponse({
        "success": False, 
        "error": "Invalid method"
    }, status=405)


# views.py
@login_required
@profile_required('hod')
def hod_dashboard(request):
    if not request.active_profile.is_hod:
        return HttpResponseForbidden()

    department = request.active_profile.department
    current_year = timezone.now().year
    year_range = range(current_year, current_year - 5, -1)

    # Update queries to use profiles instead of profile
    dept_kaizens = KaizenSheet.objects.filter(
        employee__profiles__department=department,
        employee__profiles__user_type='employee'
    )
    dept_completed = dept_kaizens.filter(approval_status="completed").count()
    dept_hod_pending = dept_kaizens.filter(approval_status="pending").count()
    dept_coordinator_pending = dept_kaizens.filter(
        approval_status="coordinator_pending"
    ).count()
    dept_finance_pending = dept_kaizens.filter(
        approval_status="finance_pending"
    ).count()

    # Get department employees
    department_employees = User.objects.filter(
        profiles__department=department,
        profiles__user_type='employee'
    )

    horizontal_sheets = (
        KaizenSheet.objects.filter(deployments__department=department)
        .exclude(employee__profiles__department=department)
        .distinct()
        .order_by("-created_at")
    )

    kaizen_sheets = KaizenSheet.objects.filter(
        employee__profiles__department=department,
        employee__profiles__user_type='employee'
    ).order_by('-created_at')



    for sheet in kaizen_sheets:
        # Add property for HOD approval button visibility
        sheet.needs_hod_approval = not sheet.hod_approved
        # Add property for approval flow
        sheet.needs_coordinator_approval = sheet.approval_status == 'coordinator_pending'
        sheet.needs_finance_approval = sheet.approval_status == 'finance_pending'

    context = {
        "kaizen_sheets": dept_kaizens.order_by("-created_at"),
        "horizontal_sheets": horizontal_sheets,
        "dept_completed": dept_completed,
        "dept_hod_pending": dept_hod_pending,
        "dept_coordinator_pending": dept_coordinator_pending,
        "dept_finance_pending": dept_finance_pending,
        "department_employees": department_employees,
        "year_range": year_range,
    }

    return render(request, "dashboard/hod_dashboard.html", context)


# Add new view for employee submissions data
@login_required
def get_employee_submissions(request, employee_id):
    # Update permission check to allow coordinators
    if not (request.active_profile.is_coordinator or request.active_profile.is_hod):
        return JsonResponse({"error": "Unauthorized"}, status=403)

    try:
        employee = get_object_or_404(User, id=employee_id)
        selected_year = request.GET.get("year", timezone.now().year)
        months = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]
        monthly_data = [0] * 12

        # Get submissions for selected year
        submissions = KaizenSheet.objects.filter(
            employee=employee, created_at__year=selected_year
        )

        # Count submissions by month
        for submission in submissions:
            month_idx = submission.created_at.month - 1
            monthly_data[month_idx] += 1

        print(
            f"Employee {employee.username} data for {selected_year}:",
            monthly_data,
        )  # Debug log

        return JsonResponse({"submissions": monthly_data, "months": months})

    except Exception as e:
        print(f"Error in get_employee_submissions: {e}")  # Debug log
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def upload_handwritten_sheet(request):
    if request.method == "POST" and request.POST.get("form_type") == "handwritten":
        try:
            title = request.POST.get("handwritten_title")
            handwritten_sheet = request.FILES.get("handwritten_sheet")

            if not title or not handwritten_sheet:
                error_msg = []
                if not title:
                    error_msg.append("Title is required")
                if not handwritten_sheet:
                    error_msg.append("Handwritten sheet is required")
                return JsonResponse({"success": False, "error": " and ".join(error_msg)})

            # Create kaizen sheet for handwritten submission
            kaizen = KaizenSheet.objects.create(
                title=title,
                employee=request.user,
                is_handwritten=True,
                handwritten_sheet=handwritten_sheet,
                area_implemented=request.active_profile.department,  # Use employee's department
                start_date=timezone.now().date(),
                end_date=timezone.now().date(),
                problem="See handwritten sheet", 
                idea_solved="See handwritten sheet",
                standardization="See handwritten sheet",
                deployment="See handwritten sheet",
                approval_status="pending"  # Start with HOD approval pending
            )

            return JsonResponse({
                "success": True,
                "message": "Handwritten sheet uploaded successfully",
                "serial_key": kaizen.serial_key
            })

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


@login_required
@profile_required('finance')
def finance_dashboard(request):
    if not request.active_profile.user_type == "finance":
        return HttpResponseForbidden()

    # Only get kaizens that need finance approval (>1 lakh)
    high_value_kaizens = KaizenSheet.objects.filter(
        models.Q(cost_before_implementation__gt=100000)
        | models.Q(cost_after_implementation__gt=100000)
    ).order_by("-created_at")

    context = {
        "kaizens": high_value_kaizens,
        "total_count": high_value_kaizens.count(),
        "pending_count": high_value_kaizens.filter(
            approval_status="finance_pending"
        ).count(),
        "approved_count": high_value_kaizens.filter(
            models.Q(approval_status="completed")
            | models.Q(approval_status="coordinator_pending")
        ).count(),
    }

    return render(request, "dashboard/finance_dashboard.html", context)

@login_required
def admin_dashboard(request):
    # Check if user is admin
    if not request.user.profiles.filter(user_type='admin').exists():
        return HttpResponseForbidden("Access denied. Admin privileges required.")
        
    context = {
        'departments': [dept[0] for dept in Profile.DEPARTMENT_CHOICES],
        'users': User.objects.exclude(
            profiles__user_type='admin'
        ).prefetch_related('profiles'),
        'password_requests': PasswordResetRequest.objects.all().order_by('-request_date'),
    }
    return render(request, 'dashboard/admin_dashboard.html', context)


@login_required
def delete_user(request):
    if not request.user.profiles.filter(user_type='admin').exists():
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        user_type = data.get('user_type')
        
        if not user_id or not user_type:
            return JsonResponse({
                'success': False, 
                'error': 'Missing required data'
            }, status=400)
            
        with transaction.atomic():
            user = User.objects.get(id=user_id)
            profile = user.profiles.filter(user_type=user_type).first()
            
            if profile:
                if user_type == 'employee':
                    # Anonymize kaizen sheets
                    KaizenSheet.objects.filter(employee=user).update(
                        employee_info={
                            'username': 'Account Deleted',
                            'employee_id': 'Deleted',
                            'department': profile.department
                        }
                    )
                    
                    # Clear department association
                    profile.department = None
                    profile.employee_id = 'Deleted'
                    profile.save()
                    
                # Delete the profile
                profile.delete()
                
                # If no more profiles exist, anonymize and deactivate user
                if not user.profiles.exists():
                    user.username = f"deleted_user_{user.id}"
                    user.email = f"deleted_{user.id}@deleted.com"
                    user.is_active = False
                    user.save()
                
                return JsonResponse({'success': True})
            else:
                return JsonResponse({
                    'success': False, 
                    'error': 'Profile not found'
                }, status=404)
            
    except User.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'User not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=500)
    
@login_required
def add_department(request):
    if not request.user.profiles.filter(user_type='admin').exists():
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    
    try:
        data = json.loads(request.body)
        department = data.get('department')
        
        if not department:
            return JsonResponse({
                'success': False, 
                'error': 'Department name required'
            }, status=400)
        
        # Check if department exists
        existing_depts = [dept[0] for dept in Profile.DEPARTMENT_CHOICES]
        if department in existing_depts:
            return JsonResponse({
                'success': False, 
                'error': 'Department already exists'
            }, status=400)
            
        # Add department
        Profile.DEPARTMENT_CHOICES = list(Profile.DEPARTMENT_CHOICES)
        Profile.DEPARTMENT_CHOICES.append((department, department))
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=500)

@login_required
def delete_department(request):
    if not request.user.profiles.filter(user_type='admin').exists():
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    
    try:
        data = json.loads(request.body)
        department = data.get('department')
        
        if not department:
            return JsonResponse({
                'success': False, 
                'error': 'Department name required'
            }, status=400)
        
        # Remove department
        Profile.DEPARTMENT_CHOICES = [
            dept for dept in Profile.DEPARTMENT_CHOICES 
            if dept[0] != department
        ]
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=500)
    
def forgot_password(request):
    if request.method == 'POST':
        try:
            employee_id = request.POST.get('employee_id')
            email = request.POST.get('email')
            user_type = request.POST.get('user_type')

            if not all([employee_id, email, user_type]):
                return JsonResponse({
                    'success': False,
                    'error': 'All fields are required'
                })

            # Verify user exists with given employee_id and user_type
            profile = Profile.objects.filter(
                employee_id=employee_id,
                user_type=user_type
            ).first()

            if not profile:
                return JsonResponse({
                    'success': False,
                    'error': 'No account found with these credentials'
                })

            # Create password reset request
            PasswordResetRequest.objects.create(
                employee_id=employee_id,
                email=email,
                department=profile.department or 'N/A',  # Use department from profile if exists
                status='pending'
            )

            return JsonResponse({
                'success': True,
                'message': 'Password reset request submitted successfully'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def generate_random_password(length=12):
    """Generate a random password with specified length"""
    # Define character sets
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits
    special = "!@#$%^&*"
    
    # Ensure at least one of each type
    password = [
        random.choice(lowercase),
        random.choice(uppercase),
        random.choice(digits),
        random.choice(special)
    ]
    
    # Fill rest with random characters
    remaining_length = length - len(password)
    all_chars = lowercase + uppercase + digits + special
    password.extend(random.choice(all_chars) for _ in range(remaining_length))
    
    # Shuffle the password
    random.shuffle(password)
    return ''.join(password)

@login_required
def handle_password_reset(request):
    if not request.user.profiles.filter(user_type='admin').exists():
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            request_id = data.get('request_id')
            action = data.get('action')

            reset_request = PasswordResetRequest.objects.get(id=request_id)
            
            if action == 'approve':
                # Find user by employee ID
                profile = Profile.objects.filter(
                    employee_id=reset_request.employee_id
                ).first()
                
                if not profile:
                    return JsonResponse({
                        'success': False,
                        'error': 'User profile not found'
                    })
                
                user = profile.user
                
                # Generate new password using our custom function
                new_password = generate_random_password()
                user.set_password(new_password)
                user.save()
                
                reset_request.status = 'approved'
                reset_request.resolved_by = request.user
                reset_request.resolved_date = timezone.now()
                reset_request.save()

                return JsonResponse({
                    'success': True,
                    'username': user.username,
                    'password': new_password,
                    'message': 'Request approved successfully'
                })
                
            elif action == 'reject':
                reset_request.status = 'rejected'
                reset_request.resolved_by = request.user
                reset_request.resolved_date = timezone.now()
                reset_request.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Request rejected successfully'
                })
            
            return JsonResponse({
                'success': False,
                'error': 'Invalid action'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def check_password_age(request):
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Not authenticated'})
    
    try:
        # Get the active profile
        active_profile = request.user.profiles.filter(id=request.session.get('active_profile_id')).first()
        
        if not active_profile:
            return JsonResponse({'success': False, 'error': 'No active profile found'})
        
        last_password_change = active_profile.last_password_change
        
        if not last_password_change:
            # If no password change recorded, use user's date_joined
            last_password_change = request.user.date_joined
        
        days_since_change = (timezone.now() - last_password_change).days
        should_change_password = days_since_change >= 40
        
        return JsonResponse({
            'success': True,
            'should_change_password': should_change_password,
            'days_since_change': days_since_change
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})